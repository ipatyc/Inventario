# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import csv
import io
import unicodedata
import openpyxl
import zipfile
import re
import datetime
import difflib
from difflib import SequenceMatcher
from openpyxl.styles import Font, PatternFill, Alignment

# ================= 1. CONFIGURACIÓN Y FUNCIONES ESTRUCTURALES =================
HOJA_ALTAS = "ALTAS"
HOJA_SALIDA_NRC = "NRC"  
UMBRAL_FUZZY = 0.82  

# Formato puro para Oracle Banner
CSV_KWARGS_R = {
    'index': False,
    'encoding': 'utf-8',
    'sep': ',',
    'lineterminator': '\n'
}

# Plantilla estricta de 24 columnas para el Clúster
COLUMNAS_CLUSTER_FINAL = [
    "Periodo", "CRN", "Tipo.de.Reunión", "Fecha.Inicio", "Fecha.Fin", "Dom", "Lun", 
    "Mar", "Mie", "Jue", "Vie", "Sab", "horarioIni", "horarioFin", "Inicio.de.sesión", 
    "edificio", "salon", "Tipo.de.horario", "indCategoria", "idInstructor", 
    "responsabilidad", "Ind.principal", "ind.sobre.paso", "datocomplementario"
]

def quitar_acentos(t):
    if pd.isna(t) or t is None: 
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(t)) if unicodedata.category(c) != "Mn")

def normalizar_para_cruce(t):
    if pd.isna(t) or t is None:
        return ""
    s = str(t).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return quitar_acentos(s).upper().strip()

def similitud(a, b): 
    return SequenceMatcher(None, a, b).ratio()

def limpiar_clave_texto(val):
    if pd.isna(val) or val is None:
        return ""
    s = str(val).strip()
    if s.lower() == "nan" or s == "":
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s

def format_r_string(val):
    if pd.isna(val) or val is None:
        return np.nan
    s = str(val).strip()
    if s.lower() == "nan" or s == "":
        return np.nan
    if s.endswith(".0"): 
        s = s[:-2]
    return s

def limpia_seccion_interna(x):
    if pd.isna(x): 
        return ""
    s = str(x).strip()
    if s.lower() == "nan" or s == "": 
        return ""
    if s.endswith(".0"): 
        s = s[:-2]
    if s.isdigit(): 
        return f"{int(s):02d}"
    return s

def sin_espacios(x):
    v = format_r_string(x)
    if pd.isna(v): return ""
    return str(v).replace(" ", "").upper()

def limpiar_espacios_y_mayusculas(x):
    if pd.isna(x): return ""
    s = re.sub(r'\s+', ' ', str(x))
    return s.strip().upper()

def normalizar_para_busqueda(texto):
    s = str(texto).lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r'[^a-z0-9]', '', s)

def limpiar_nombre_columna(col):
    if pd.isna(col): return ""
    return " ".join(str(col).split())

def normalizar_para_busqueda_t3(texto):
    s = str(texto).lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r'[^a-z0-9]', '', s)

def ultra_limpiar(x):
    if pd.isna(x): return ""
    s = str(x).strip().upper().replace(" ", "")
    if s.endswith(".0"): s = s[:-2]
    return s

def ultra_limpiar_seccion(x):
    if pd.isna(x): return ""
    s = str(x).strip().upper().replace(" ", "")
    if s.endswith(".0"): s = s[:-2]
    if s.isdigit(): return f"{int(s):02d}"
    return s

def corregir_nivel_por_cluster_csv(cluster_val):
    cluster_str = str(cluster_val).strip().lower()
    if "posgrado" in cluster_str: return "POSGRADO"
    elif "bachillerato" in cluster_str: return "BACHILLERATO"
    else: return "LICENCIATURA"

def simplificar_nombre(nombre):
    n = nombre.lower()
    for basura in ['.xlsx', '.xls', '.csv', '_final', '_base', '_v1', '_v2', '_v3', '_v4', 'corregidas_', 'errores_']:
        n = n.replace(basura, '')
    return n.strip().replace(" ", "")

# Inicialización de estados en memoria global
if "original_files_bytes" not in st.session_state: st.session_state.original_files_bytes = {}
if "res_auditoria" not in st.session_state: st.session_state.res_auditoria = None
if "raw_altas" not in st.session_state: st.session_state.raw_altas = None
if "ready_for_download" not in st.session_state: st.session_state.ready_for_download = False
if "zip_file_bytes" not in st.session_state: st.session_state.zip_file_bytes = None
if "csv_files_to_download" not in st.session_state: st.session_state.csv_files_to_download = {}
if "delta_files" not in st.session_state: st.session_state.delta_files = {}
if "final_argos_zip" not in st.session_state: st.session_state.final_argos_zip = None
if "df_cruce_rapido" not in st.session_state: st.session_state.df_cruce_rapido = None
if "df_delta_cache" not in st.session_state: st.session_state.df_delta_cache = None
if "nombre_delta_cache" not in st.session_state: st.session_state.nombre_delta_cache = None
if "llave_control_archivos" not in st.session_state: st.session_state.llave_control_archivos = ""
if "archivo_final_bytes" not in st.session_state: st.session_state.archivo_final_bytes = None
if "archivo_final_nombre" not in st.session_state: st.session_state.archivo_final_nombre = None

st.set_page_config(page_title="Consola Iris Cavazos", page_icon="⚙️", layout="wide")
st.title("⚙️ Consola de Control de Materias e Inyección de NRCs")
st.markdown("---")

tab1, tab_err, tab3 = st.tabs([
    "1️⃣ Proceso: Validación y Generar CSVs", 
    "⚠️ Reporte de Errores (Extraer Delta)", 
    "2️⃣ Proceso: Inyección de NRCs Masiva (ARGOS)"
])

# ============================================================
# PESTAÑA 1: VALIDACIÓN, CSV Y AUTOCOMPLETADO
# ============================================================
with tab1:
    # --- 1. ENCABEZADO Y BOTÓN DE REINICIO ---
    col_tit, col_btn = st.columns([4, 1])
    with col_tit:
        st.header("Validación de Claves, Horarios y Generación de CSV")
    with col_btn:
        # Limpia todas las variables de sesión para empezar desde cero
        if st.button("🔄 Limpiar / Recomenzar", type="secondary", use_container_width=True, key="btn_limpiar_t1"):
            claves_a_borrar = [
                "file_cat_uploader", "file_cat_ext_uploader", "files_altas_uploader", "res_auditoria", 
                "raw_altas", "ready_for_download", "modo_salida_csv", "zip_file_bytes", 
                "csv_consolidado_bytes", "modo_salida_csv_generado", "cat_avanzado_cache", 
                "cat_avanzado_firma", "df_manual_fijo", "manual_candidatos", "manual_busqueda_realizada", 
                "manual_materia_seleccionada", "manual_archivo_visualizado", "manual_renglon_accion", 
                "input_nom_busq", "input_subj_busq", "input_crse_busq", "manual_horario", "manual_metodo", 
                "manual_periodo", "manual_parte_periodo", "manual_capacidad", "manual_seccion", 
                "manual_sede", "manual_estatus", "manual_modo_calificar", "manual_sesion", "manual_grupos", 
                "manual_integracion", "manual_nivel", "manual_cluster", "manual_materia_confirmada"
            ]
            for clave in claves_a_borrar:
                if clave in st.session_state: del st.session_state[clave]
            st.rerun()

    # --- 2. ZONA DE CARGA DE ARCHIVOS ---
    col1, col2, col3 = st.columns(3)
    with col1: file_cat = st.file_uploader("📑 Catálogo Básico (Niveles)", type=["xlsx"], key="file_cat_uploader")
    with col2: file_cat_ext = st.file_uploader("📚 Catálogo Avanzado (SCBCRSE)", type=["csv", "xlsx"], key="file_cat_ext_uploader")
    with col3: files_altas = st.file_uploader("📁 Archivos ALTAS", accept_multiple_files=True, type=["xlsx"], key="files_altas_uploader")

    # Configuración de salida de CSV si hay archivos cargados
    modo_salida_csv = "Un CSV por cada Excel"
    if files_altas:
        modo_salida_csv = st.radio("¿Cómo deseas generar los CSV de los archivos ALTAS?", ["Un CSV por cada Excel", "Un solo CSV consolidado"], horizontal=True, key="modo_salida_csv")

    columnas_esperadas = ["Periodo", "Campus", "Subject", "Course", "Nivel", "Nombre de la Materia", "Parte de Periodo", "Estatus", "Capacidad", "Sección", "Tipo de Horario", "Método Educativo", "Modo de Calificar", "Sesion", "Clúster"]
    mapa_huellas = {normalizar_para_busqueda(col): col for col in columnas_esperadas}

    # Valores permitidos para Dato Complementario / Clúster.
    CLUSTERS_PERMITIDOS = [
        "Ingenieria", "Bachillerato", "Negocios", "Ciencias Exactas",
        "Posgrado Online", "Humanidades", "Idiomas y ADN", "TJYG",
        "Smart Cities", "Ejecutivas", "EGEL",
        "Intercambio", "Consejeria", "Posgrado"
    ]
    MAPA_CLUSTERS = {
        normalizar_para_busqueda(cluster): cluster
        for cluster in CLUSTERS_PERMITIDOS
    }

    UMBRAL_SIMILITUD_CLUSTER = 0.90

    def obtener_cluster_permitido(valor):
        # Corrige solo clusters muy parecidos; los demás se conservan como llegaron.
        cluster_original = format_r_string(valor)
        cluster_normalizado = normalizar_para_busqueda(cluster_original)

        if cluster_normalizado in MAPA_CLUSTERS:
            return MAPA_CLUSTERS[cluster_normalizado]

        mejor_cluster, mejor_similitud = cluster_original, 0.0
        for huella_cluster, cluster_oficial in MAPA_CLUSTERS.items():
            coincidencia = similitud(cluster_normalizado, huella_cluster)
            if coincidencia > mejor_similitud:
                mejor_cluster, mejor_similitud = cluster_oficial, coincidencia

        return mejor_cluster if mejor_similitud >= UMBRAL_SIMILITUD_CLUSTER else cluster_original

    # --- 3. FUNCIÓN PARA CARGAR CATÁLOGO AVANZADO EN MEMORIA ---
    def cargar_catalogo_avanzado():
        firma_catalogo = f"{file_cat_ext.name}:{getattr(file_cat_ext, 'size', '')}" if file_cat_ext else None
        if "cat_avanzado_cache" in st.session_state and st.session_state.get("cat_avanzado_firma") == firma_catalogo:
            return st.session_state.cat_avanzado_cache, st.session_state.indice_nombres_avanzado

        cat_avanzado, indice_nombres_avanzado = {}, {}
        if file_cat_ext is not None:
            try:
                df_ext = pd.read_csv(file_cat_ext, dtype=str, encoding="utf-8", on_bad_lines="skip") if file_cat_ext.name.lower().endswith(".csv") else pd.read_excel(file_cat_ext, dtype=str)
                df_ext.columns = [str(col).strip().upper() for col in df_ext.columns]

                for _, row in df_ext.iterrows():
                    subj, crse = sin_espacios(row.get("SCBCRSE_SUBJ_CODE")), sin_espacios(row.get("SCBCRSE_CRSE_NUMB"))
                    if not subj or not crse: continue

                    titulo_corto, titulo_largo = limpiar_espacios_y_mayusculas(row.get("SCBCRSE_TITLE")), limpiar_espacios_y_mayusculas(row.get("SCRSYLN_LONG_COURSE_TITLE"))
                    horario, metodo = sin_espacios(row.get("SCRSCHD_SCHD_CODE")), sin_espacios(row.get("SCRSCHD_INSM_CODE"))
                    modo_calificar = sin_espacios(row.get("SCRGMOD_GMOD_CODE"))
                    llave = (subj, crse)

                    # AHORA GUARDAMOS LAS PAREJAS EXACTAS DE HORARIO-MÉTODO
                    if llave not in cat_avanzado:
                        cat_avanzado[llave] = {"titles": set(), "schd": set(), "insm": set(), "gmod": set(), "pares": set()}

                    if titulo_corto and titulo_corto != "NAN":
                        cat_avanzado[llave]["titles"].add(titulo_corto)
                        indice_nombres_avanzado[normalizar_para_cruce(titulo_corto)] = llave
                    if titulo_largo and titulo_largo != "NAN":
                        cat_avanzado[llave]["titles"].add(titulo_largo)
                        indice_nombres_avanzado[normalizar_para_cruce(titulo_largo)] = llave
                    
                    h_val = horario if horario and horario != "NAN" else ""
                    m_val = metodo if metodo and metodo != "NAN" else ""
                    
                    if h_val: cat_avanzado[llave]["schd"].add(h_val)
                    if m_val: cat_avanzado[llave]["insm"].add(m_val)
                    if modo_calificar and modo_calificar != "NAN": cat_avanzado[llave]["gmod"].add(modo_calificar)
                    if h_val or m_val: cat_avanzado[llave]["pares"].add((h_val, m_val))

                st.session_state.cat_avanzado_cache, st.session_state.indice_nombres_avanzado = cat_avanzado, indice_nombres_avanzado
                st.session_state.cat_avanzado_firma = firma_catalogo
            except Exception as error:
                st.error(f"Error al leer el Catálogo Avanzado: {error}")
        return cat_avanzado, indice_nombres_avanzado

    # --- 4. VALIDACIÓN DE ARCHIVOS MASIVOS (ALTAS) ---
    if files_altas and file_cat:
        if st.button("⚡ Ejecutar Validación Inteligente", type="primary", key="btn_val_inteligente"):
            st.session_state.ready_for_download = False
            st.toast("Cargando Catálogos y validando...", icon="📑")
            
            cat_avanzado, indice_nombres_avanzado = cargar_catalogo_avanzado()
            xls_cat = pd.ExcelFile(file_cat)
            indice_cat, indice_cat_claves = {}, {}

            # Indexar catálogo básico
            for hoja in xls_cat.sheet_names:
                df_catalogo = xls_cat.parse(hoja)
                if "Nivel" in df_catalogo.columns and "Materia" in df_catalogo.columns:
                    for _, fila in df_catalogo.iterrows():
                        nivel, materia, subj, crse = normalizar_para_cruce(fila.get("Nivel")), limpiar_espacios_y_mayusculas(fila.get("Materia")), sin_espacios(fila.get("Subj")), sin_espacios(fila.get("Crse"))
                        indice_cat.setdefault(nivel, []).append({"mat_orig": materia, "mat_norm": normalizar_para_cruce(fila.get("Materia")), "subj": subj, "crse": crse})
                        if subj and crse: indice_cat_claves[(normalizar_para_cruce(subj), crse)] = materia

            # Busca la coincidencia más cercana usando nombre, SUBJ y COURSE.
            def buscar_mejor_coincidencia(materia, subj, crse, nivel):
                candidatos = []

                for (subj_cat, crse_cat), info in cat_avanzado.items():
                    titulos = sorted(info["titles"]) or [""]
                    candidatos.append({
                        "subj": subj_cat, "crse": crse_cat, "materia": titulos[0],
                        "titulos": titulos, "fuente": "Cat. Avanzado"
                    })

                basicos_nivel = indice_cat.get(nivel, [])
                basicos = basicos_nivel or [fila for filas in indice_cat.values() for fila in filas]
                for candidato in basicos:
                    candidatos.append({
                        "subj": candidato["subj"], "crse": candidato["crse"],
                        "materia": candidato["mat_orig"], "titulos": [candidato["mat_orig"]],
                        "fuente": "Cat. Básico"
                    })

                mejor, mejor_puntaje = None, -1.0
                for candidato in candidatos:
                    puntaje_nombre = max(
                        [similitud(materia, normalizar_para_cruce(titulo)) for titulo in candidato["titulos"]] or [0.0]
                    )
                    puntaje_subj = 1.0 if subj == candidato["subj"] else similitud(subj, candidato["subj"])
                    puntaje_crse = 1.0 if crse == candidato["crse"] else similitud(crse, candidato["crse"])
                    puntaje = (puntaje_nombre * 0.55) + (puntaje_subj * 0.25) + (puntaje_crse * 0.20)

                    if puntaje > mejor_puntaje:
                        mejor, mejor_puntaje = candidato, puntaje

                return mejor, mejor_puntaje

            # Procesar archivos ALTAS
            piezas, resultados = [], []
            for archivo_altas in files_altas:
                st.session_state.original_files_bytes[archivo_altas.name] = archivo_altas.getvalue()
                xls_altas = pd.ExcelFile(archivo_altas)
                hojas_reales = [h for h in xls_altas.sheet_names if h.strip().upper() == HOJA_ALTAS]
                if not hojas_reales: continue

                df_altas = xls_altas.parse(hojas_reales[0], dtype=str)
                df_altas.columns = [mapa_huellas.get(normalizar_para_busqueda(col), col) for col in df_altas.columns]
                
                cols_esenciales = [c for c in ["Periodo", "Campus", "Subject", "Course"] if c in df_altas.columns]
                if cols_esenciales: df_altas = df_altas.dropna(subset=cols_esenciales, how="all").dropna(how="all")
                
                if not df_altas.empty:
                    df_altas["ArchivoOrigen"] = archivo_altas.name
                    piezas.append(df_altas)

            if piezas:
                df_total = pd.concat(piezas, ignore_index=True)
                st.session_state.raw_altas = df_total.copy()

                # Validación fila por fila
                for idx, fila in df_total.iterrows():
                    nivel, materia_excel, subj_original, crse_original = normalizar_para_cruce(fila.get("Nivel")), limpiar_espacios_y_mayusculas(fila.get("Nombre de la Materia")), sin_espacios(fila.get("Subject")), sin_espacios(fila.get("Course"))
                    horario_original, metodo_original = sin_espacios(fila.get("Tipo de Horario")), sin_espacios(fila.get("Método Educativo"))
                    modo_calificar_original = sin_espacios(fila.get("Modo de Calificar"))
                    materia_normalizada = normalizar_para_cruce(materia_excel)
                    
                    subj_sugerido, crse_sugerido, materia_catalogo, comentario_nombres = subj_original, crse_original, materia_excel, ""

                    # Validación Nombres y Claves: compara contra ambos catálogos.
                    mejor_candidato, mejor_puntaje = buscar_mejor_coincidencia(
                        materia_normalizada, subj_original, crse_original, nivel
                    )

                    if mejor_candidato and mejor_puntaje >= 0.63:
                        subj_sugerido = mejor_candidato["subj"]
                        crse_sugerido = mejor_candidato["crse"]
                        materia_catalogo = mejor_candidato["materia"]
                        mismo_subj_crse = subj_original == subj_sugerido and crse_original == crse_sugerido
                        mismo_nombre = materia_normalizada in [
                            normalizar_para_cruce(titulo) for titulo in mejor_candidato["titulos"]
                        ]

                        if mismo_subj_crse and mismo_nombre:
                            comentario_nombres = f"Todo correcto ({mejor_candidato['fuente']})"
                        elif mismo_subj_crse:
                            comentario_nombres = f"Clave OK, pero Nombre difiere ({mejor_candidato['fuente']})"
                        else:
                            comentario_nombres = (
                                f"Claves sugeridas ({mejor_candidato['fuente']}, "
                                f"similitud {mejor_puntaje:.0%})"
                            )
                    else:
                        comentario_nombres = "No se encontró una coincidencia suficiente en ningún catálogo"

                    # Validación Horarios, Métodos y Modo de Calificar
                    horario_sugerido, metodo_sugerido = horario_original, metodo_original
                    modo_calificar_sugerido = modo_calificar_original
                    comentario_horario, comentario_metodo = "Sin catálogo para validar", "Sin catálogo para validar"
                    comentario_modo_calificar = "Sin catálogo para validar"
                    
                    if cat_avanzado and (subj_sugerido, crse_sugerido) in cat_avanzado:
                        horarios_permitidos = cat_avanzado[(subj_sugerido, crse_sugerido)]["schd"]
                        metodos_permitidos = cat_avanzado[(subj_sugerido, crse_sugerido)]["insm"]
                        modos_calificar_permitidos = cat_avanzado[(subj_sugerido, crse_sugerido)]["gmod"]
                        
                        if horarios_permitidos:
                            if horario_original in horarios_permitidos: comentario_horario = "Horario OK"
                            else:
                                comentario_horario = f"Error. Permitidos: {', '.join(sorted(horarios_permitidos))}"
                                horario_sugerido = sorted(horarios_permitidos)[0] if len(horarios_permitidos) == 1 else ""
                        else: comentario_horario = "Sin restricciones en catálogo"

                        if metodos_permitidos:
                            if metodo_original in metodos_permitidos: comentario_metodo = "Método OK"
                            else:
                                comentario_metodo = f"Error. Permitidos: {', '.join(sorted(metodos_permitidos))}"
                                metodo_sugerido = sorted(metodos_permitidos)[0] if len(metodos_permitidos) == 1 else ""
                        else: comentario_metodo = "Sin restricciones en catálogo"

                        if modos_calificar_permitidos:
                            if modo_calificar_original in modos_calificar_permitidos: comentario_modo_calificar = "Modo de calificar OK"
                            else:
                                comentario_modo_calificar = f"Error. Permitidos: {', '.join(sorted(modos_calificar_permitidos))}"
                                modo_calificar_sugerido = sorted(modos_calificar_permitidos)[0] if len(modos_calificar_permitidos) == 1 else ""
                        else: comentario_modo_calificar = "Sin restricciones en catálogo"

                    resultados.append({
                        "Luz Verde": False, "idx": idx, "Archivo": fila.get("ArchivoOrigen"),
                        "Materia Excel": materia_excel, "Materia Catálogo": materia_catalogo, "Comentario Nombres": comentario_nombres,
                        "Subj Original": subj_original, "Crse Original": crse_original, "Subj Sugerido": subj_sugerido, "Crse Sugerido": crse_sugerido,
                        "Horario Original": horario_original, "Horario Sugerido": horario_sugerido, "Comentario Horario": comentario_horario,
                        "Método Original": metodo_original, "Método Sugerido": metodo_sugerido, "Comentario Método": comentario_metodo,
                        "Modo de Calificar Original": modo_calificar_original, "Modo de Calificar Sugerido": modo_calificar_sugerido,
                        "Comentario Modo de Calificar": comentario_modo_calificar,
                        "Llave_Cruce": f"{fila.get('ArchivoOrigen')}|{materia_excel}|{subj_original}|{crse_original}|{idx}"
                    })

                st.session_state.res_auditoria = pd.DataFrame(resultados)
                st.success("¡Revisión de catálogos finalizada!")
            else:
                st.error(f"❌ Ninguno de los archivos subidos tiene filas válidas en la pestaña '{HOJA_ALTAS}'.")

    # --- 5. MESA DE CONTROL (REVISIÓN DE ERRORES) ---
    if st.session_state.res_auditoria is not None:
        st.markdown("### ⚖️ Mesa de Control (Dividida en 2 Partes)")
        df_auditoria = st.session_state.res_auditoria
        
        for archivo in df_auditoria["Archivo"].unique():
            df_archivo = df_auditoria[df_auditoria["Archivo"] == archivo]
            filas_con_error = df_archivo[
                (~df_archivo["Comentario Nombres"].str.startswith("Todo correcto", na=False)) |
                (~df_archivo["Comentario Horario"].isin(["Horario OK", "Sin restricciones en catálogo", "Sin catálogo para validar"])) |
                (~df_archivo["Comentario Método"].isin(["Método OK", "Sin restricciones en catálogo", "Sin catálogo para validar"])) |
                (~df_archivo["Comentario Modo de Calificar"].isin(["Modo de calificar OK", "Sin restricciones en catálogo", "Sin catálogo para validar"]))
            ]
            
            if filas_con_error.empty:
                st.success(f"✅ **{archivo}** — Claves, horarios, métodos y modo de calificar validados.")
            else:
                with st.expander(f"⚠️ **{archivo}** — ({len(filas_con_error)} advertencias detectadas)", expanded=True):
                    with st.form(key=f"form_{archivo}"):
                        col_nombres, col_metodos = st.tabs(["PARTE 1: Nombres y Claves", "PARTE 2: Métodos y Horarios"])
                        with col_nombres:
                            df_editado_nombres = st.data_editor(filas_con_error[["Luz Verde", "Materia Excel", "Materia Catálogo", "Comentario Nombres", "Subj Original", "Crse Original", "Subj Sugerido", "Crse Sugerido"]], hide_index=True, disabled=["Materia Excel", "Materia Catálogo", "Comentario Nombres", "Subj Original", "Crse Original"], column_config={"Luz Verde": st.column_config.CheckboxColumn("¿Aplicar?")}, key=f"edit_nom_{archivo}", use_container_width=True)
                        with col_metodos:
                            df_editado_metodos = st.data_editor(
                                filas_con_error[[
                                    "Luz Verde", "Materia Excel", "Horario Original", "Horario Sugerido", "Comentario Horario",
                                    "Método Original", "Método Sugerido", "Comentario Método", "Modo de Calificar Original",
                                    "Modo de Calificar Sugerido", "Comentario Modo de Calificar"
                                ]],
                                hide_index=True,
                                disabled=[
                                    "Materia Excel", "Horario Original", "Comentario Horario", "Método Original", "Comentario Método",
                                    "Modo de Calificar Original", "Comentario Modo de Calificar"
                                ],
                                column_config={"Luz Verde": st.column_config.CheckboxColumn("¿Aplicar?")},
                                key=f"edit_met_{archivo}", use_container_width=True
                            )
                        
                        if st.form_submit_button("💾 Confirmar Selección de Ambas Pestañas"):
                            df_final_edits = filas_con_error.copy()
                            df_final_edits["Luz Verde"] = df_editado_nombres["Luz Verde"] | df_editado_metodos["Luz Verde"]
                            df_final_edits[["Subj Sugerido", "Crse Sugerido"]] = df_editado_nombres[["Subj Sugerido", "Crse Sugerido"]]
                            df_final_edits[["Horario Sugerido", "Método Sugerido", "Modo de Calificar Sugerido"]] = df_editado_metodos[["Horario Sugerido", "Método Sugerido", "Modo de Calificar Sugerido"]]
                            
                            df_master = st.session_state.res_auditoria.copy().set_index("Llave_Cruce")
                            df_final_edits.set_index("Llave_Cruce", inplace=True)
                            df_master.update(df_final_edits[["Luz Verde", "Subj Sugerido", "Crse Sugerido", "Horario Sugerido", "Método Sugerido", "Modo de Calificar Sugerido"]])
                            st.session_state.res_auditoria = df_master.reset_index()
                            st.rerun()

        # --- 6. GENERADOR MASIVO DE CSV ---
        if st.button("💾 Generar Bloque de Archivos CSV", type="primary", key="btn_generar_bloque_csv"):
            st.session_state.ready_for_download = False
            corregido = st.session_state.raw_altas.copy()
            for col in ["Subject", "Course", "Tipo de Horario", "Método Educativo", "Modo de Calificar"]: 
                if col in corregido.columns: corregido[col] = corregido[col].astype(str)

            for _, fila in st.session_state.res_auditoria[st.session_state.res_auditoria["Luz Verde"]].iterrows():
                if pd.notna(fila["Subj Sugerido"]): corregido.loc[fila["idx"], "Subject"] = str(fila["Subj Sugerido"])
                if pd.notna(fila["Crse Sugerido"]): corregido.loc[fila["idx"], "Course"] = str(fila["Crse Sugerido"])
                if pd.notna(fila["Horario Sugerido"]) and fila["Horario Sugerido"] != "": corregido.loc[fila["idx"], "Tipo de Horario"] = str(fila["Horario Sugerido"])
                if pd.notna(fila["Método Sugerido"]) and fila["Método Sugerido"] != "": corregido.loc[fila["idx"], "Método Educativo"] = str(fila["Método Sugerido"])
                if pd.notna(fila["Modo de Calificar Sugerido"]) and fila["Modo de Calificar Sugerido"] != "": corregido.loc[fila["idx"], "Modo de Calificar"] = str(fila["Modo de Calificar Sugerido"])

            def preparar_csv_banner(df_origen):
                res = pd.DataFrame()
                res["PERIODO"] = df_origen["Periodo"].apply(format_r_string)
                res["SEDE"] = df_origen["Campus"].apply(format_r_string)
                res["SUBJ"] = df_origen["Subject"].apply(sin_espacios)
                res["COURSE"] = df_origen["Course"].apply(sin_espacios)
                res["PARTEPERIODO"] = df_origen["Parte de Periodo"].apply(format_r_string)
                res["STATUS"] = df_origen["Estatus"].apply(format_r_string)
                res["CAPACIDAD"] = pd.to_numeric(df_origen["Capacidad"], errors="coerce").astype("Int64")
                res["GRUPOS"] = pd.Series(1, index=res.index, dtype="Int64")
                res["SECCION"] = pd.to_numeric(df_origen["Sección"], errors="coerce").astype("Int64")
                res["TIPODEHORARIO"] = df_origen["Tipo de Horario"].apply(sin_espacios)
                res["METODO_EDUCATIVO"] = df_origen["Método Educativo"].apply(sin_espacios)
                res["SOCIODEINTEGRACION"] = "D2L"
                res["MODODECALIFICAR"] = df_origen["Modo de Calificar"].apply(format_r_string)
                res["SESION"] = df_origen["Sesion"].apply(format_r_string)

                def aplicar_reglas_cluster(fila):
                    nivel_actual = str(fila.get("Nivel", "")).strip().upper()
                    cluster_excel = obtener_cluster_permitido(fila.get("Clúster"))

                    if "BACHILLERATO" in nivel_actual:
                        return "Bachillerato"

                    return cluster_excel

                res["datocomplementario"] = df_origen.apply(aplicar_reglas_cluster, axis=1)
                
                # Limpiar texto
                for col in res.columns: res[col] = res[col].astype(str).str.replace('"', "", regex=False).str.strip().replace(["nan", "None", "<NA>", "NaN"], "")
                return res[["PERIODO", "SEDE", "SUBJ", "COURSE", "PARTEPERIODO", "STATUS", "CAPACIDAD", "GRUPOS", "SECCION", "TIPODEHORARIO", "METODO_EDUCATIVO", "SOCIODEINTEGRACION", "MODODECALIFICAR", "SESION", "datocomplementario"]].to_csv(**CSV_KWARGS_R)

            st.session_state.csv_files_to_download, st.session_state.zip_file_bytes, st.session_state.csv_consolidado_bytes = {}, None, None
            errores_encontrados = False
            columnas_requeridas_csv = ["Periodo", "Campus", "Subject", "Course", "Nivel", "Parte de Periodo", "Estatus", "Capacidad", "Sección", "Tipo de Horario", "Método Educativo", "Modo de Calificar", "Sesion", "Clúster"]

            if modo_salida_csv == "Un CSV por cada Excel":
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    for nom_arch, sub in corregido.groupby("ArchivoOrigen"):
                        faltantes = [c for c in columnas_requeridas_csv if c not in sub.columns]
                        if faltantes:
                            st.error(f"❌ Error en `{nom_arch}`. Faltan columnas: {', '.join(faltantes)}."); errores_encontrados = True; continue
                        csv_string = preparar_csv_banner(sub)
                        nombre_csv = f"{nom_arch.rsplit('.', 1)[0]}.csv" if "." in nom_arch else f"{nom_arch}.csv"
                        zip_file.writestr(nombre_csv, csv_string.encode("utf-8"))
                        st.session_state.csv_files_to_download[nombre_csv] = csv_string.encode("utf-8")
                if not errores_encontrados: st.session_state.zip_file_bytes = zip_buffer.getvalue()
            else:
                for nom_arch, sub in corregido.groupby("ArchivoOrigen"):
                    faltantes = [c for c in columnas_requeridas_csv if c not in sub.columns]
                    if faltantes: st.error(f"❌ Faltan columnas en `{nom_arch}`: {', '.join(faltantes)}"); errores_encontrados = True
                if not errores_encontrados: st.session_state.csv_consolidado_bytes = preparar_csv_banner(corregido).encode("utf-8")

            if not errores_encontrados:
                st.session_state.ready_for_download = True
                st.session_state.modo_salida_csv_generado = modo_salida_csv
                st.rerun()

        # Panel de descargas Masivo
        if st.session_state.ready_for_download:
            st.markdown("### 📥 Panel de Descarga")
            modo_desc = st.session_state.modo_salida_csv_generado or modo_salida_csv
            if modo_desc == "Un CSV por cada Excel":
                st.download_button("📥 Descargar todos los CSV (.ZIP)", data=st.session_state.zip_file_bytes, file_name="archivos_carga_banner.zip", mime="application/zip", use_container_width=True, type="primary", key="dl_todos_csv_zip")
            else:
                st.download_button("📥 Descargar CSV consolidado", data=st.session_state.csv_consolidado_bytes, file_name="archivos_carga_banner.csv", mime="text/csv", use_container_width=True, type="primary", key="dl_csv_consolidado")

    # ============================================================
    # --- 7. MÓDULO MANUAL (BUSCADOR FLEXIBLE CON FORMULARIO) ---
    # ============================================================
    st.markdown("---")
    tiene_archivos_altas = bool(files_altas)
    
    st.subheader("Visualizador de ALTAS y agregar registro" if tiene_archivos_altas else "Creación de CSV Manual")
    st.caption("Agrega materias sueltas. Si buscas, el sistema autocompletará y filtrará las opciones de horario y método.")

    if "df_manual_fijo" not in st.session_state: st.session_state.df_manual_fijo = pd.DataFrame(columns=["PERIODO", "SEDE", "SUBJ", "COURSE", "PARTEPERIODO", "STATUS", "CAPACIDAD", "GRUPOS", "SECCION", "TIPODEHORARIO", "METODO_EDUCATIVO", "SOCIODEINTEGRACION", "MODODECALIFICAR", "SESION", "datocomplementario"])
    if "manual_candidatos" not in st.session_state: st.session_state.manual_candidatos = []
    if "manual_busqueda_realizada" not in st.session_state: st.session_state.manual_busqueda_realizada = False

    archivo_destino_manual = None
    if tiene_archivos_altas:
        if st.session_state.raw_altas is None: st.info("Para agregar registros al Excel, primero ejecuta la Validación Inteligente arriba.")
        else:
            archivos_disp = sorted(st.session_state.raw_altas["ArchivoOrigen"].dropna().unique().tolist())
            if archivos_disp and st.session_state.get("manual_archivo_visualizado") not in archivos_disp: st.session_state.manual_archivo_visualizado = archivos_disp[0]
            archivo_destino_manual = st.selectbox("Excel que deseas visualizar y completar", options=archivos_disp, key="manual_archivo_visualizado")

    def buscar_candidatos_manual(cat_avanzado, nombre, subj, crse, limite=20):
        nombre_norm, subj_norm, crse_norm = normalizar_para_busqueda(nombre) if nombre else "", normalizar_para_busqueda(subj) if subj else "", normalizar_para_busqueda(crse) if crse else ""
        resultados = []
        for (subj_cat, crse_cat), info in cat_avanzado.items():
            subj_cat_norm, crse_cat_norm, puntaje, campos = normalizar_para_busqueda(subj_cat), normalizar_para_busqueda(crse_cat), 0.0, 0
            if subj_norm:
                sim = max(1.0 if subj_norm == subj_cat_norm else similitud(subj_norm, subj_cat_norm), 0.95 if subj_norm in subj_cat_norm or subj_cat_norm in subj_norm else 0.0)
                if sim < 0.55: continue
                puntaje += sim; campos += 1
            if crse_norm:
                sim = max(1.0 if crse_norm == crse_cat_norm else similitud(crse_norm, crse_cat_norm), 0.95 if crse_norm in crse_cat_norm or crse_cat_norm in crse_norm else 0.0)
                if sim < 0.55: continue
                puntaje += sim; campos += 1
            
            titulos = sorted(info["titles"])
            titulo_elegido = titulos[0] if titulos else "SIN TÍTULO"
            if nombre_norm:
                sim = max([similitud(nombre_norm, normalizar_para_busqueda(t)) for t in titulos] or [0.0])
                if any(nombre_norm in normalizar_para_busqueda(t) for t in titulos): sim = max(sim, 0.95)
                if sim < 0.40 and not subj_norm and not crse_norm: continue
                puntaje += sim; campos += 1
                
            if campos: resultados.append({"llave": (subj_cat, crse_cat), "titulo": titulo_elegido, "puntaje": puntaje / campos})
        return sorted(resultados, key=lambda d: (-d["puntaje"], d["llave"]))[:limite]

    # ST.FORM: ESTO EVITA QUE SE PIERDA EL TEXTO AL ESCRIBIR
    st.markdown("#### 🔍 Buscador de materia")
    with st.form("form_buscador_manual"):
        col_busq1, col_busq2, col_busq3 = st.columns(3)
        input_nombre_busq = col_busq1.text_input("Nombre o título", key="input_nom_busq").strip()
        input_subj_busq = col_busq2.text_input("SUBJ", key="input_subj_busq").strip()
        input_crse_busq = col_busq3.text_input("COURSE", key="input_crse_busq").strip()
        
        btn_buscar_manual = st.form_submit_button("🪄 Buscar opciones", type="secondary", use_container_width=True)

    if btn_buscar_manual:
        if not file_cat_ext: st.warning("Primero sube el Catálogo Avanzado para buscar.")
        elif not input_nombre_busq and not input_subj_busq and not input_crse_busq: st.warning("Ingresa al menos un criterio (nombre, SUBJ o COURSE).")
        else:
            cat_avanzado, _ = cargar_catalogo_avanzado()
            st.session_state.manual_candidatos = buscar_candidatos_manual(cat_avanzado, input_nombre_busq, input_subj_busq, input_crse_busq)
            st.session_state.manual_busqueda_realizada = True
            st.session_state.pop("manual_materia_seleccionada", None)

    # --- 8. ZONA PARA AGREGAR LA MATERIA BUSCADA ---
    candidatos = st.session_state.manual_candidatos
    if candidatos:
        opc_mat = [c["llave"] for c in candidatos]
        etiq_mat = {c["llave"]: f"{c['llave'][0]} {c['llave'][1]} - {c['titulo']} ({c['puntaje']:.0%})" for c in candidatos}
        if st.session_state.get("manual_materia_seleccionada") not in opc_mat: st.session_state.manual_materia_seleccionada = opc_mat[0]
        
        llave_materia = st.selectbox("Selecciona la materia correcta", options=opc_mat, format_func=lambda k: etiq_mat[k], key="manual_materia_seleccionada")
        confirmar_materia = st.checkbox(
            f"Confirmo que quiero agregar: {etiq_mat[llave_materia]}",
            value=False,
            key=f"confirmar_materia_{llave_materia[0]}_{llave_materia[1]}"
        )
        
        cat_avanzado, _ = cargar_catalogo_avanzado()
        info_materia = cat_avanzado[llave_materia]
        
        # === LÓGICA DE FILTRADO DEPENDIENTE (HORARIOS <-> MÉTODOS) ===
        horarios_base = sorted(info_materia["schd"])
        metodos_base = sorted(info_materia["insm"])
        modos_calificar_base = [""] + sorted(info_materia.get("gmod", set()))
        pares_validos = info_materia.get("pares", set())

        sel_horario = st.session_state.get("manual_horario", "")
        sel_metodo = st.session_state.get("manual_metodo", "")

        # Filtramos horarios si ya hay un método seleccionado
        if sel_metodo and sel_metodo in metodos_base:
            horarios_validos = sorted({h for h, m in pares_validos if m == sel_metodo and h})
        else:
            horarios_validos = horarios_base

        # Filtramos métodos si ya hay un horario seleccionado
        if sel_horario and sel_horario in horarios_base:
            metodos_validos = sorted({m for h, m in pares_validos if h == sel_horario and m})
        else:
            metodos_validos = metodos_base

        horarios_disp = [""] + horarios_validos
        metodos_disp = [""] + metodos_validos

        # Evitar errores de Streamlit si la opción guardada ya no es válida tras el filtro
        if sel_horario not in horarios_disp: st.session_state.manual_horario = ""
        if sel_metodo not in metodos_disp: st.session_state.manual_metodo = ""
        # ==============================================================

        st.markdown("#### Datos para agregar la materia")
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            horario_manual = st.selectbox("Tipo de horario", options=horarios_disp, key="manual_horario")
            periodo_manual = st.text_input("Periodo", key="manual_periodo").strip()
            parte_periodo_manual = st.text_input("Parte de periodo", key="manual_parte_periodo").strip()
            capacidad_manual = st.text_input("Capacidad", key="manual_capacidad").strip()
            seccion_manual = st.text_input("Sección", key="manual_seccion").strip()
        with col_d2:
            metodo_manual = st.selectbox("Método educativo", options=metodos_disp, key="manual_metodo")
            sede_manual = st.text_input("Sede", key="manual_sede").strip()
            estatus_manual = st.text_input("Estatus", key="manual_estatus").strip()
            modo_calificar_manual = st.selectbox(
                "Modo de calificar",
                options=modos_calificar_base,
                key="manual_modo_calificar",
                # NOTA: st.selectbox no tiene "accept_new_options" en la librería estándar. 
                # Si usas una librería externa cámbialo, o reemplázalo por text_input si el usuario debe poder escribir
            ).strip()
            sesion_manual = st.text_input("Sesión", key="manual_sesion").strip()

        col_e1, col_e2, col_e3, col_extra4 = st.columns(4)
        with col_e1: grupos_manual = st.number_input("Grupos", min_value=1, value=1, step=1, key="manual_grupos")
        with col_e2: nivel_manual = st.selectbox("Nivel", options=["LICENCIATURA", "BACHILLERATO", "POSGRADO"], key="manual_nivel")
        with col_e3: integracion_manual = st.text_input("Socio de integración", value="D2L", key="manual_integracion").strip()
        with col_extra4: 
            cluster_manual = st.selectbox(
                "Dato Complementario (Clúster)",
                options=CLUSTERS_PERMITIDOS,
                index=None,
                placeholder="Selecciona un Clúster",
                key="manual_cluster"
            )

        if st.button("Agregar materia seleccionada", type="primary", use_container_width=True, key="btn_agregar_manual"):
            if not confirmar_materia:
                st.warning("Confirma la materia seleccionada antes de agregarla. Así no se carga por accidente ni se modifica tu revisión.")
                st.stop()
            if not cluster_manual:
                st.warning("Selecciona un Dato Complementario / Clúster antes de agregar la materia.")
                st.stop()
                
            nuevo_renglon_manual = pd.DataFrame([{
                "PERIODO": periodo_manual, "SEDE": sede_manual, "SUBJ": llave_materia[0], "COURSE": llave_materia[1],
                "PARTEPERIODO": parte_periodo_manual, "STATUS": estatus_manual, "CAPACIDAD": capacidad_manual,
                "GRUPOS": str(grupos_manual), "SECCION": seccion_manual, "TIPODEHORARIO": horario_manual,
                "METODO_EDUCATIVO": metodo_manual, "SOCIODEINTEGRACION": integracion_manual or "D2L",
                "MODODECALIFICAR": modo_calificar_manual, "SESION": sesion_manual,
                "datocomplementario": "Bachillerato" if nivel_manual == "BACHILLERATO" else cluster_manual
            }])

            if tiene_archivos_altas:
                if st.session_state.raw_altas is None or not archivo_destino_manual: st.warning("Primero ejecuta la validación y selecciona el Excel destino.")
                else:
                    titulo_mat = next(c["titulo"] for c in candidatos if c["llave"] == llave_materia)
                    nuevo_renglon_excel = {col: "" for col in st.session_state.raw_altas.columns}
                    nuevo_renglon_excel.update({
                        "Periodo": periodo_manual, "Campus": sede_manual, "Subject": llave_materia[0], "Course": llave_materia[1],
                        "Nivel": nivel_manual, "Nombre de la Materia": titulo_mat, "Parte de Periodo": parte_periodo_manual,
                        "Estatus": estatus_manual, "Capacidad": capacidad_manual, "Sección": seccion_manual,
                        "Tipo de Horario": horario_manual, "Método Educativo": metodo_manual, "Modo de Calificar": modo_calificar_manual,
                        "Sesion": sesion_manual, "Clúster": cluster_manual, "ArchivoOrigen": archivo_destino_manual
                    })
                    st.session_state.raw_altas = pd.concat([st.session_state.raw_altas, pd.DataFrame([nuevo_renglon_excel])], ignore_index=True)
                    st.success(f"Materia agregada al Excel `{archivo_destino_manual}`.")
            else:
                st.session_state.df_manual_fijo = pd.concat([st.session_state.df_manual_fijo, nuevo_renglon_manual], ignore_index=True)
                st.success("Materia agregada a la tabla manual.")

    elif st.session_state.manual_busqueda_realizada: st.info("No se encontraron coincidencias con esos datos.")

    # --- 9. VISUALIZADOR DE TABLAS (EXCEL O MANUAL) ---
    if tiene_archivos_altas and st.session_state.raw_altas is not None and archivo_destino_manual:
        st.markdown("#### Vista previa del Excel seleccionado")
        cols_vis = [c for c in ["Periodo", "Campus", "Subject", "Course", "Nivel", "Nombre de la Materia", "Parte de Periodo", "Estatus", "Capacidad", "Sección", "Tipo de Horario", "Método Educativo", "Modo de Calificar", "Sesion", "Clúster"] if c in st.session_state.raw_altas.columns]
        st.dataframe(st.session_state.raw_altas.loc[st.session_state.raw_altas["ArchivoOrigen"] == archivo_destino_manual, cols_vis], hide_index=True, use_container_width=True)

    # ============================================================
    # TABLA MANUAL: SELECCIONAR, COPIAR, ELIMINAR O AGREGAR RENGLONES
    # ============================================================
    if not tiene_archivos_altas:
        st.markdown("#### Tabla manual")

        # La columna Seleccionar solo sirve para acciones; no se exporta al CSV.
        df_editor_manual = st.session_state.df_manual_fijo.copy()
        if "Seleccionar" not in df_editor_manual.columns:
            df_editor_manual.insert(0, "Seleccionar", False)

        df_editado = st.data_editor(
            df_editor_manual,
            num_rows="fixed", # Evita que se creen renglones por accidente al escribir abajo
            use_container_width=True,
            hide_index=True,
            key="editor_manual_seguro",
            column_config={
                "Seleccionar": st.column_config.CheckboxColumn(
                    "Seleccionar",
                    default=False
                )
            }
        )

        seleccionados = df_editado.index[df_editado["Seleccionar"]].tolist()

        # Guardamos los cambios de la tabla sin la columna de selección
        # BUG FIX: Comparamos antes de guardar para no perder el foco mientras escribes
        df_sin_seleccion = df_editado.drop(columns="Seleccionar")
        if not st.session_state.df_manual_fijo.equals(df_sin_seleccion):
            st.session_state.df_manual_fijo = df_sin_seleccion

        col_accion1, col_accion2, col_accion3 = st.columns(3)

        with col_accion1:
            if st.button("Nuevo renglón", use_container_width=True, key="btn_nuevo_renglon_manual"):
                # Toma automáticamente las columnas que sí existen en la tabla
                nuevo_renglon_vacio = {col: "" for col in st.session_state.df_manual_fijo.columns}
                nuevo_renglon_vacio["GRUPOS"] = "1"
                nuevo_renglon_vacio["SOCIODEINTEGRACION"] = "D2L"

                st.session_state.df_manual_fijo = pd.concat(
                    [st.session_state.df_manual_fijo, pd.DataFrame([nuevo_renglon_vacio])],
                    ignore_index=True
                )
                st.rerun()

        with col_accion2:
            if st.button("Copiar seleccionados", use_container_width=True, key="btn_copiar_renglones_manual"):
                if not seleccionados:
                    st.warning("Marca al menos un renglón en la columna 'Seleccionar'.")
                else:
                    renglones_copia = st.session_state.df_manual_fijo.loc[seleccionados].copy()
                    st.session_state.df_manual_fijo = pd.concat(
                        [st.session_state.df_manual_fijo, renglones_copia],
                        ignore_index=True
                    )
                    st.rerun()

        with col_accion3:
            if st.button("Eliminar seleccionados", use_container_width=True, key="btn_eliminar_renglones_manual"):
                if not seleccionados:
                    st.warning("Marca al menos un renglón en la columna 'Seleccionar'.")
                else:
                    st.session_state.df_manual_fijo = (
                        st.session_state.df_manual_fijo
                        .drop(index=seleccionados)
                        .reset_index(drop=True)
                    )
                    st.rerun()

        # --- 10. DESCARGA DEL ARCHIVO MANUAL ---
        col_nom, col_desc = st.columns([3, 1])
        nombre_csv_manual = col_nom.text_input("Nombre del archivo:", value="carga_manual.csv", key="nom_manual_seguro")
        
        # Al descargar usamos df_sin_seleccion para que no se incluya el checkbox
        df_out_manual = df_sin_seleccion.copy()
        
        # 1. Aplicar format_r_string a las columnas de texto generales (Igual que en el CSV masivo)
        columnas_r_string = ["PERIODO", "SEDE", "PARTEPERIODO", "STATUS", "MODODECALIFICAR", "SESION", "datocomplementario"]
        for col in columnas_r_string:
            if col in df_out_manual.columns:
                df_out_manual[col] = df_out_manual[col].apply(format_r_string)
        
        # 2. Aplicar sin_espacios a las claves (Igual que en el CSV masivo)
        for col in ["SUBJ", "COURSE", "TIPODEHORARIO", "METODO_EDUCATIVO"]: 
            if col in df_out_manual.columns: 
                df_out_manual[col] = df_out_manual[col].apply(sin_espacios)
                
        # 3. Forzar valores fijos y numéricos (Igual que en el CSV masivo)
        if "GRUPOS" in df_out_manual.columns: 
            df_out_manual["GRUPOS"] = pd.Series(1, index=df_out_manual.index, dtype="Int64")
        if "SOCIODEINTEGRACION" in df_out_manual.columns: 
            df_out_manual["SOCIODEINTEGRACION"] = "D2L"
            
        for col_num in ["CAPACIDAD", "SECCION"]:
            if col_num in df_out_manual.columns: 
                df_out_manual[col_num] = pd.to_numeric(df_out_manual[col_num], errors="coerce").astype("Int64")

        # 4. Limpieza final de texto: quitar comillas dobles y nulos de pandas
        for col in df_out_manual.columns: 
            if df_out_manual[col].dtype == object or pd.api.types.is_string_dtype(df_out_manual[col]):
                df_out_manual[col] = df_out_manual[col].astype(str).str.replace('"', "", regex=False).str.strip().replace(["nan", "None", "<NA>", "NaN"], "")
        
        # 5. Forzar el ORDEN EXACTO de las columnas
        orden_columnas = [
            "PERIODO", "SEDE", "SUBJ", "COURSE", "PARTEPERIODO", "STATUS", 
            "CAPACIDAD", "GRUPOS", "SECCION", "TIPODEHORARIO", "METODO_EDUCATIVO", 
            "SOCIODEINTEGRACION", "MODODECALIFICAR", "SESION", "datocomplementario"
        ]
        columnas_existentes = [c for c in orden_columnas if c in df_out_manual.columns]
        df_out_manual = df_out_manual[columnas_existentes]

        col_desc.write(""); col_desc.write("") # Espaciado para alinear el botón
        col_desc.download_button(
            label="📥 Descargar CSV Manual", 
            data=df_out_manual.to_csv(**CSV_KWARGS_R).encode("utf-8"),
            file_name=nombre_csv_manual if nombre_csv_manual.endswith(".csv") else f"{nombre_csv_manual}.csv",
            mime="text/csv", 
            type="primary", 
            use_container_width=True, 
            key="dl_csv_manual_btn"
        )


# ============================================================
# PESTAÑA 2: REPORTE DE ERRORES Y ENSAMBLAJE FINAL
# ============================================================
with tab_err:
    # --- ENCABEZADO Y BOTÓN DE REINICIO ---
    col_tit_t2, col_btn_t2 = st.columns([4, 1])
    with col_tit_t2:
        st.header("⚠️ Reporte de Errores y Ensamblaje Final")
    with col_btn_t2:
        # Limpia todas las variables de sesión de esta pestaña
        if st.button("🔄 Limpiar / Recomenzar", type="secondary", use_container_width=True, key="btn_limpiar_t2"):
            claves_a_borrar_t2 = [
                "df_delta_cache", "nombre_delta_cache", "llave_control_archivos", 
                "archivo_final_bytes", "archivo_final_nombre", "ext_base_1", "ext_err_1", 
                "suf_v1", "modo_1", "ed_vivo_1", "iny_base_2", "iny_err_2", "iny_corr_2", "suf_v2"
            ]
            for clave in claves_a_borrar_t2:
                if clave in st.session_state: del st.session_state[clave]
            st.rerun()
            
    st.markdown("Extrae filas con error, corrígelas y genera el archivo para la Pestaña 3.")
    
    # --- PASO 1: EXTRAER O EDITAR EL PEDACITO CON ERROR ---
    st.subheader("✂️ 1. Extraer o corregir el pedacito con errores")
    
    col_ex1, col_ex2, col_ex3 = st.columns(3)
    with col_ex1: file_base_ext = st.file_uploader("📁 1. Archivo Base (.csv)", type=["csv"], key="ext_base_1")
    with col_ex2: file_err_ext = st.file_uploader("📊 2. Reporte de Errores Banner (.xlsx)", type=["xlsx"], key="ext_err_1")
    with col_ex3: sufijo_version = st.text_input("🔢 Sufijo de versión (Ej: V1, V2):", value="V1", key="suf_v1")
    
    if file_base_ext and file_err_ext:
        llave_actual = f"{file_base_ext.name}_{file_err_ext.name}_{sufijo_version}"
        if st.session_state.llave_control_archivos != llave_actual:
            st.session_state.df_delta_cache = None
            st.session_state.nombre_delta_cache = None
            st.session_state.llave_control_archivos = llave_actual

        if st.button("🔍 Cargar y Procesar Reporte de Errores", use_container_width=True, type="secondary", key="btn_cargar_reporte_errores"):
            try:
                df_base = pd.read_csv(file_base_ext, encoding="utf-8", dtype=str)
                df_err = pd.read_excel(file_err_ext, skiprows=2)
                
                df_err.columns = [limpiar_nombre_columna(c) for c in df_err.columns]
                col_linea = [c for c in df_err.columns if "linea" in str(c).strip().lower().replace("í", "i")]
                
                if not col_linea:
                    st.error("❌ No se encontró la columna 'Línea' en el reporte de errores. Verifica la estructura de tu archivo.")
                else:
                    nombre_col = col_linea[0]
                    df_err = df_err.dropna(subset=[nombre_col])
                    
                    indices = [int(float(r)) - 2 for r in df_err[nombre_col].unique().tolist() if pd.notna(r) and 0 <= (int(float(r)) - 2) < len(df_base)]
                    
                    if indices:
                        st.session_state.df_delta_cache = df_base.iloc[indices].copy()
                        base_name_ext = file_base_ext.name.rsplit('.', 1)[0].replace("_base", "").replace("_final", "")
                        st.session_state.nombre_delta_cache = f"{base_name_ext}_{sufijo_version}"
                        st.success(f"🎉 Éxito: Se aislaron {len(indices)} filas con anomalías. Configura tu descarga abajo.")
                    else:
                        st.warning("⚠️ No se identificaron números de línea válidos dentro de las dimensiones del archivo base.")
            except Exception as e:
                st.error(f"❌ Error crítico de lectura física: {str(e)}")

    if st.session_state.df_delta_cache is not None:
        st.markdown("---")
        modo_delta = st.radio("⚙️ ¿Cómo deseas descargar o corregir el fragmento?", ["Excel (.xlsx)", "CSV (.csv)", "Editar en vivo"], horizontal=True, key="modo_1")
        
        nombre_archivo = st.session_state.nombre_delta_cache
        df_delta = st.session_state.df_delta_cache
        
        if modo_delta == "Excel (.xlsx)":
            buf = io.BytesIO()
            df_delta.to_excel(buf, index=False)
            st.download_button("📥 Descargar Fragmento (.xlsx)", data=buf.getvalue(), file_name=f"{nombre_archivo}.xlsx", type="primary", use_container_width=True, key="dl_frag_xlsx")
        
        elif modo_delta == "CSV (.csv)":
            st.download_button("📥 Descargar Fragmento (.csv)", data=df_delta.to_csv(**CSV_KWARGS_R).encode("utf-8"), file_name=f"{nombre_archivo}.csv", type="primary", use_container_width=True, key="dl_frag_csv")
        
        else:
            st.info("✏️ **Modo Edición Interactiva:** Escribe tus ajustes directamente en las celdas de la tabla. Al finalizar, haz clic en el botón inferior para exportarlo.")
            df_editado = st.data_editor(df_delta, key="ed_vivo_1", use_container_width=True)
            st.download_button("📥 Descargar Parche Corregido (.csv)", data=df_editado.to_csv(**CSV_KWARGS_R).encode("utf-8"), file_name=f"{nombre_archivo}.csv", type="primary", use_container_width=True, key="dl_parche_corr_csv")

    # --- PASO 2: INYECTAR Y CREAR EL ARCHIVO FINAL ---
    st.subheader("💉 2. Inyectar correcciones y generar Archivo Final")
    
    col_in1, col_in2, col_in3 = st.columns(3)
    with col_in1: file_base_iny = st.file_uploader("📁 1. Archivo Base (.csv)", type=["csv"], key="iny_base_2")
    with col_in2: file_err_iny = st.file_uploader("📊 2. Reporte de Errores (.xlsx)", type=["xlsx"], key="iny_err_2")
    with col_in3: 
        file_corr_iny = st.file_uploader("📝 3. Fragmento Corregido", type=["csv", "xlsx"], key="iny_corr_2")
        tipo_final = st.text_input("Etiqueta final (V1, V2, final):", value="final", key="suf_v2")
    
    if file_base_iny and file_err_iny and file_corr_iny:
        if st.button("🚀 Ensamblar Archivo Final", type="primary", key="btn_ensamblar_archivo_final"):
            try: 
                df_base = pd.read_csv(file_base_iny, encoding="utf-8", dtype=str)
                df_err = pd.read_excel(file_err_iny, skiprows=2)
                
                col_linea_iny = [c for c in df_err.columns if "linea" in str(c).strip().lower().replace("í", "i")]
                
                if not col_linea_iny:
                    st.error("❌ No se encontró la columna 'Línea' en el reporte.")
                else:
                    nombre_col_iny = col_linea_iny[0]
                    df_err = df_err.dropna(subset=[nombre_col_iny])
                    df_corr = pd.read_excel(file_corr_iny, dtype=str) if file_corr_iny.name.endswith('.xlsx') else pd.read_csv(file_corr_iny, encoding="utf-8", dtype=str)
                    
                    indices = [int(float(r)) - 2 for r in df_err[nombre_col_iny].unique().tolist() if pd.notna(r) and 0 <= (int(float(r)) - 2) < len(df_base)]
                    
                    if len(indices) == len(df_corr):
                        df_final = df_base.copy()
                        for col in df_final.columns:
                            if col in df_corr.columns: df_final.iloc[indices, df_final.columns.get_loc(col)] = df_corr[col].values
                        
                        for col in df_final.columns:
                            df_final[col] = df_final[col].astype(str).str.replace('"', '', regex=False).str.strip().replace(['nan', 'None', '<NA>', 'NaN'], '')
                        
                        base_name_iny = file_base_iny.name.rsplit('.', 1)[0].replace("_base", "").replace("_final", "")
                        out_name = f"{base_name_iny}_{tipo_final}.csv"
                        
                        st.session_state.archivo_final_bytes = df_final.to_csv(**CSV_KWARGS_R).encode("utf-8")
                        st.session_state.archivo_final_nombre = out_name
                        st.success(f"🎉 ¡Archivo {out_name} listo! Da clic en el botón debajo para descargar.")
                    else:
                        st.error(f"❌ Desajuste de filas: {len(indices)} errores en Banner vs {len(df_corr)} filas corregidas en tu archivo.")
            except Exception as e:
                st.error(f"❌ Error interno al ensamblar: {str(e)}")

    if st.session_state.archivo_final_bytes is not None:
        st.download_button(
            label=f"📁 📥 DESCARGAR {st.session_state.archivo_final_nombre}", 
            data=st.session_state.archivo_final_bytes, 
            file_name=st.session_state.archivo_final_nombre, 
            type="primary", 
            use_container_width=True,
            key="dl_archivo_final_ensamblado"
        )
# ============================================================
# PESTAÑA 3: INYECCIÓN DE NRCS Y CRUCES CON ARGOS
# ============================================================
with tab3:
    # --- ENCABEZADO Y BOTÓN DE REINICIO ---
    col_tit_t3, col_btn_t3 = st.columns([4, 1])
    with col_tit_t3:
        st.header("Inyección de NRCs y Cruces con ARGOS")
    with col_btn_t3:
        # Limpia todas las variables de sesión de esta pestaña
        if st.button("🔄 Limpiar / Recomenzar", type="secondary", use_container_width=True, key="btn_limpiar_t3"):
            claves_a_borrar_t3 = [
                "modo_inyeccion_t3", "arg_c", "csv_c", "xls_c", "final_argos_zip", 
                "arg_r", "csv_r", "df_cruce_rapido", "columnas_copia_rapida"
            ]
            for clave in claves_a_borrar_t3:
                if clave in st.session_state: del st.session_state[clave]
            st.rerun()
            
    modo_inyeccion = st.radio(
        "🛠️ **Elige tu escenario de archivos disponibles:**",
        ["📦 Completo (Tengo ARGOS, CSV Final y Excel Original)", 
         "⚡ Rápido (Tengo ARGOS y CSV Final)"],
        horizontal=True,
        key="modo_inyeccion_t3"
    )
    st.markdown("---")

    columnas_esperadas_t3 = [
        "Periodo", "Campus", "Subject", "Course", "Nivel", "Nombre de la Materia",
        "Parte de Periodo", "Estatus", "Capacidad", "Sección", 
        "Tipo de Horario", "Método Educativo", "Modo de Calificar", "Sesion", "Clúster"
    ]
    mapa_huellas_t3 = {normalizar_para_busqueda_t3(col): col for col in columnas_esperadas_t3}

    # ====================================================================
    # MODO A: COMPLETO (3 ARCHIVOS)
    # ====================================================================
    if modo_inyeccion == "📦 Completo (Tengo ARGOS, CSV Final y Excel Original)":
        st.markdown("Procesa los Excels originales, inyecta los NRCs de ARGOS cruzando con el CSV final y genera un ZIP con los Excels actualizados.")
        col_a, col_b, col_c = st.columns(3)
        with col_a: file_argos = st.file_uploader("📊 1. Reporte ARGOS (.csv)", type=["csv"], key="arg_c")
        with col_b: files_csv_finales = st.file_uploader("📝 2. CSVs Finales", type=["csv"], accept_multiple_files=True, key="csv_c")
        with col_c: files_xlsx_originales = st.file_uploader("📁 3. Excels Originales", type=["xlsx"], accept_multiple_files=True, key="xls_c")
            
        if file_argos and files_csv_finales and files_xlsx_originales:
            if st.button("🚀 PROCESAR Y GENERAR EXCELS CON NRC", type="primary", key="btn_procesar_argos_completo"):
                try:
                    argos_df = pd.read_csv(file_argos, encoding="utf-8", on_bad_lines='skip', dtype=str)
                    argos_df.columns = [re.sub(r'\.+', '.', str(c).replace('"', '').replace("'", "").strip()) for c in argos_df.columns]
                    
                    col_argos_cluster = next((c for c in argos_df.columns if "cluster" in normalizar_para_busqueda_t3(c)), None)
                    col_argos_area = next((c for c in argos_df.columns if "area" in normalizar_para_busqueda_t3(c)), None)
                    col_argos_curso = next((c for c in argos_df.columns if "curso" in normalizar_para_busqueda_t3(c)), None)

                    if not col_argos_cluster: raise KeyError("No se encontró la columna de Cluster en ARGOS.")
                    if not col_argos_area: raise KeyError("No se encontró la columna de Área en ARGOS.")
                    if not col_argos_curso: raise KeyError("No se encontró la columna de Curso en ARGOS.")

                    argos_df["Periodo"] = argos_df["Periodo"].apply(ultra_limpiar)
                    argos_df["Nivel"] = argos_df["Nivel"].apply(ultra_limpiar)
                    argos_df[col_argos_cluster] = argos_df[col_argos_cluster].apply(ultra_limpiar)
                    argos_df[col_argos_area] = argos_df[col_argos_area].apply(ultra_limpiar)
                    argos_df[col_argos_curso] = argos_df[col_argos_curso].apply(ultra_limpiar)
                    argos_df["Grupo"] = argos_df["Grupo"].apply(ultra_limpiar_seccion)
                    
                    argos_df["_llave_argos"] = (
                        argos_df["Periodo"] + "_" + 
                        argos_df["Nivel"] + "_" + 
                        argos_df[col_argos_cluster] + "_" + 
                        argos_df[col_argos_area] + "_" + 
                        argos_df[col_argos_curso] + "_" + 
                        argos_df["Grupo"]
                    )
                    argos_df = argos_df.drop_duplicates(subset=["_llave_argos"])
                    mapa_nrcs = dict(zip(argos_df["_llave_argos"], argos_df["NRC"]))
                    llaves_argos_disponibles = list(mapa_nrcs.keys())

                    excels_inyectados_zip = io.BytesIO()
                    archivos_procesados_con_exito = 0
                    alertas_dimensiones, alertas_parejas = [], []
                    alertas_nrc_faltantes = []
                    
                    with zipfile.ZipFile(excels_inyectados_zip, "w", zipfile.ZIP_DEFLATED) as zip_out:
                        for fx in files_xlsx_originales:
                            df_csv, fc_usado = None, None
                            base_excel = simplificar_nombre(fx.name)
                            
                            for fc_cand in files_csv_finales:
                                base_csv = simplificar_nombre(fc_cand.name)
                                if base_excel == base_csv or base_excel in base_csv or base_csv in base_excel:
                                    df_csv = pd.read_csv(io.BytesIO(fc_cand.getvalue()), encoding="utf-8", dtype=str)
                                    fc_usado = fc_cand
                                    break
                            
                            if df_csv is not None:
                                wb = openpyxl.load_workbook(io.BytesIO(fx.getvalue()))
                                if HOJA_ALTAS in wb.sheetnames:
                                    data = list(wb[HOJA_ALTAS].values)
                                    if not data: continue
                                    
                                    df_excel_original = pd.DataFrame(data[1:], columns=[str(c).strip() if c is not None else "" for c in data[0]])
                                    
                                    nuevas_columnas = []
                                    for col in df_excel_original.columns:
                                        huella = normalizar_para_busqueda_t3(col)
                                        if huella in mapa_huellas_t3:
                                            nuevas_columnas.append(mapa_huellas_t3[huella])
                                        else:
                                            nuevas_columnas.append(col)
                                    df_excel_original.columns = nuevas_columnas
                                    
                                    df_excel_original = df_excel_original.dropna(how='all')
                                    df_csv = df_csv.dropna(how='all')
                                    
                                    if "Periodo" in df_excel_original.columns:
                                        df_excel_original = df_excel_original[df_excel_original["Periodo"].astype(str).str.strip() != ""]
                                    if "PERIODO" in df_csv.columns:
                                        df_csv = df_csv[df_csv["PERIODO"].astype(str).str.strip() != ""]
                                    
                                    df_excel_original, df_csv = df_excel_original.reset_index(drop=True), df_csv.reset_index(drop=True)
                                    
                                    if len(df_excel_original) != len(df_csv):
                                        alertas_dimensiones.append(f"❌ Excel `{fx.name}` tiene **{len(df_excel_original)} filas**, CSV `{fc_usado.name}` tiene **{len(df_csv)} filas**.")
                                        continue
                                    
                                    df_nrc_pestana = df_excel_original.copy()
                                    mapeo_columnas = {
                                        "Periodo": "PERIODO", "Campus": "SEDE", "Subject": "SUBJ", "Course": "COURSE",
                                        "Parte de Periodo": "PARTEPERIODO", "Estatus": "STATUS", "Capacidad": "CAPACIDAD",
                                        "Sección": "SECCION", "Tipo de Horario": "TIPODEHORARIO", "Método Educativo": "METODO_EDUCATIVO",
                                        "Modo de Calificar": "MODODECALIFICAR", "Sesion": "SESION"
                                    }
                                    
                                    for col_ex, col_cs in mapeo_columnas.items():
                                        if col_ex in df_nrc_pestana.columns and col_cs in df_csv.columns:
                                            if col_ex == "Sección": df_nrc_pestana[col_ex] = pd.to_numeric(df_csv[col_cs], errors='coerce').values
                                            else: df_nrc_pestana[col_ex] = df_csv[col_cs].values
                                    
                                    df_nrc_pestana["Grupos"], df_nrc_pestana["Socio de Integración"] = "1", "D2L"
                                    
                                    cluster_csv_series = df_csv["datocomplementario"] if "datocomplementario" in df_csv.columns else pd.Series([""] * len(df_csv))
                                    nivel_corregido = cluster_csv_series.apply(corregir_nivel_por_cluster_csv).apply(ultra_limpiar)
                                    cluster_limpio_csv = cluster_csv_series.apply(ultra_limpiar)
                                    
                                    llaves_cruce = (
                                        df_nrc_pestana["Periodo"].apply(ultra_limpiar) + "_" + 
                                        nivel_corregido + "_" + 
                                        cluster_limpio_csv + "_" + 
                                        df_nrc_pestana["Subject"].apply(ultra_limpiar) + "_" + 
                                        df_nrc_pestana["Course"].apply(ultra_limpiar) + "_" + 
                                        df_nrc_pestana["Sección"].apply(ultra_limpiar_seccion)
                                    )
                                    
                                    nrc_mapeados = llaves_cruce.map(mapa_nrcs)
                                    
                                    faltantes = llaves_cruce[nrc_mapeados.isna()]
                                    if not faltantes.empty:
                                        for llave_rota in faltantes.unique():
                                            sugerencias = difflib.get_close_matches(str(llave_rota), llaves_argos_disponibles, n=1, cutoff=0.5)
                                            sugerencia_txt = f"👉 En ARGOS lo más parecido es: **{sugerencias[0]}**" if sugerencias else "👉 (No se encontró nada parecido en ARGOS)"
                                            alertas_nrc_faltantes.append(f"❌ `{fx.name}` buscó: **{llave_rota}** \n{sugerencia_txt}")

                                    df_nrc_pestana.insert(0, "NRC", nrc_mapeados)
                                    
                                    if HOJA_SALIDA_NRC in wb.sheetnames: del wb[HOJA_SALIDA_NRC]
                                    ws_nrc = wb.create_sheet(title=HOJA_SALIDA_NRC)
                                    ws_nrc.append(list(df_nrc_pestana.columns))
                                    for fila in df_nrc_pestana.values: ws_nrc.append([None if pd.isna(v) else v for v in fila])
                                    
                                    font_base, font_nrc, font_header = Font(name="Calibri", size=11), Font(name="Calibri", size=11, bold=True), Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                                    fill_header, fill_nrc = PatternFill(start_color="1F4E78", fill_type="solid"), PatternFill(start_color="DDEBF7", fill_type="solid")
                                    align_header, align_center = Alignment(horizontal="center", vertical="center", wrap_text=True), Alignment(horizontal="center", vertical="center")
                                    
                                    for row in ws_nrc.iter_rows(min_row=1, max_row=ws_nrc.max_row, min_col=1, max_col=ws_nrc.max_column):
                                        for cell in row: cell.font = font_base
                                    for cell in ws_nrc[1]: cell.font, cell.fill, cell.alignment = font_header, fill_header, align_header
                                    for cell in ws_nrc['A'][1:]: cell.font, cell.fill, cell.alignment = font_nrc, fill_nrc, align_center
                                    
                                    for col in ws_nrc.columns:
                                        max_len = 0
                                        for cell in col:
                                            if cell.value: max_len = max(max_len, len(str(cell.value)))
                                        ws_nrc.column_dimensions[col[0].column_letter].width = max(max_len + 3, 11)
                                    
                                    nombre_salida_excel = fc_usado.name.rsplit('.', 1)[0] + "_con_NRC.xlsx"
                                    excel_buffer = io.BytesIO()
                                    wb.save(excel_buffer)
                                    zip_out.writestr(nombre_salida_excel, excel_buffer.getvalue())
                                    archivos_procesados_con_exito += 1
                                    
                            else:
                                alertas_parejas.append(f"⚠️ `{fx.name}` no encontró ningún CSV compatible.")

                    if archivos_procesados_con_exito > 0:
                        st.session_state.final_argos_zip = excels_inyectados_zip.getvalue()
                        st.success(f"🎉 ¡Proceso finalizado! Se procesaron {archivos_procesados_con_exito} archivos exitosamente.")
                    else: 
                        st.error("❌ No se pudo procesar ningún archivo.")
                    
                    if alertas_nrc_faltantes:
                        st.markdown("### 🔍 Radar de Llaves Rotas (Comparación Frente a Frente):")
                        for alerta in alertas_nrc_faltantes: st.error(alerta)
                            
                    if alertas_dimensiones:
                        st.markdown("### 🚫 Archivos descartados por diferencia de filas:")
                        for alerta in alertas_dimensiones: st.error(alerta)
                    if alertas_parejas:
                        st.markdown("### ❓ Archivos sin pareja:")
                        for alerta in alertas_parejas: st.warning(alerta)

                except Exception as e:
                    st.error(f"❌ Ocurrió un inconveniente crítico: {str(e)}")

        if st.session_state.final_argos_zip is not None:
            st.markdown("### 📥 Panel de Descarga (Excels Inyectados)")
            st.download_button(
                label="📁 📥 DESCARGAR EXCELS CON NRC (.ZIP)", data=st.session_state.final_argos_zip,
                file_name="Excels_Finales_con_NRC.zip", mime="application/zip",
                use_container_width=True, type="primary",
                key="dl_excels_nrc_zip"
            )

    # ====================================================================
    # MODO B: RÁPIDO (SOLO CSV FINAL + ARGOS)
    # ====================================================================
    elif modo_inyeccion == "⚡ Rápido (Tengo ARGOS y CSV Final)":
        st.markdown("Extrae los NRC de ARGOS cruzando con el CSV final y genera una tabla limpia que puedes copiar o descargar en Excel.")
        
        col_r1, col_r2 = st.columns(2)
        with col_r1: file_argos_rap = st.file_uploader("📊 1. Reporte ARGOS (.csv)", type=["csv"], key="arg_r")
        with col_r2: files_csv_rap = st.file_uploader("📝 2. CSVs Finales", type=["csv"], accept_multiple_files=True, key="csv_r")
        
        if file_argos_rap and files_csv_rap:
            if st.button("⚡ Cruzar NRC y Generar Tabla", type="primary", key="btn_cruzar_rapido"):
                try:
                    argos_df = pd.read_csv(file_argos_rap, encoding="utf-8", on_bad_lines='skip', dtype=str)
                    argos_df.columns = [re.sub(r'\.+', '.', str(c).replace('"', '').replace("'", "").strip()) for c in argos_df.columns]
                    
                    col_argos_cluster = next((c for c in argos_df.columns if "cluster" in normalizar_para_busqueda_t3(c)), None)
                    col_argos_area = next((c for c in argos_df.columns if "area" in normalizar_para_busqueda_t3(c)), None)
                    col_argos_curso = next((c for c in argos_df.columns if "curso" in normalizar_para_busqueda_t3(c)), None)

                    if not col_argos_cluster: raise KeyError("No se encontró la columna de Cluster en ARGOS.")
                    if not col_argos_area: raise KeyError("No se encontró la columna de Área en ARGOS.")
                    if not col_argos_curso: raise KeyError("No se encontró la columna de Curso en ARGOS.")

                    argos_df["Periodo"] = argos_df["Periodo"].apply(ultra_limpiar)
                    argos_df["Nivel"] = argos_df["Nivel"].apply(ultra_limpiar)
                    argos_df[col_argos_cluster] = argos_df[col_argos_cluster].apply(ultra_limpiar)
                    argos_df[col_argos_area] = argos_df[col_argos_area].apply(ultra_limpiar)
                    argos_df[col_argos_curso] = argos_df[col_argos_curso].apply(ultra_limpiar)
                    argos_df["Grupo"] = argos_df["Grupo"].apply(ultra_limpiar_seccion)
                    
                    argos_df["_llave_argos"] = (
                        argos_df["Periodo"] + "_" + 
                        argos_df["Nivel"] + "_" + 
                        argos_df[col_argos_cluster] + "_" + 
                        argos_df[col_argos_area] + "_" + 
                        argos_df[col_argos_curso] + "_" + 
                        argos_df["Grupo"]
                    )
                    argos_df = argos_df.drop_duplicates(subset=["_llave_argos"])
                    mapa_nrcs = dict(zip(argos_df["_llave_argos"], argos_df["NRC"]))
                    llaves_argos_disponibles = list(mapa_nrcs.keys())
                    
                    dfs_combinados = []
                    alertas_nrc_rapido = []
                    
                    for fc in files_csv_rap:
                        df_c = pd.read_csv(io.BytesIO(fc.getvalue()), encoding="utf-8", dtype=str)
                        df_c = df_c.dropna(how='all')
                        
                        cluster_csv_series = df_c["datocomplementario"] if "datocomplementario" in df_c.columns else pd.Series([""] * len(df_c))
                        nivel_csv = cluster_csv_series.apply(corregir_nivel_por_cluster_csv).apply(ultra_limpiar)
                        cluster_limpio_csv = cluster_csv_series.apply(ultra_limpiar)
                        
                        llaves_csv = (
                            df_c.get("PERIODO", pd.Series(dtype=str)).apply(ultra_limpiar) + "_" + 
                            nivel_csv + "_" + 
                            cluster_limpio_csv + "_" + 
                            df_c.get("SUBJ", pd.Series(dtype=str)).apply(ultra_limpiar) + "_" + 
                            df_c.get("COURSE", pd.Series(dtype=str)).apply(ultra_limpiar) + "_" + 
                            df_c.get("SECCION", pd.Series(dtype=str)).apply(ultra_limpiar_seccion)
                        )
                        
                        nrc_asignados = llaves_csv.map(mapa_nrcs)
                        
                        faltantes = llaves_csv[nrc_asignados.isna()]
                        if not faltantes.empty:
                            for llave_rota in faltantes.unique():
                                sugerencias = difflib.get_close_matches(str(llave_rota), llaves_argos_disponibles, n=1, cutoff=0.5)
                                sugerencia_txt = f"👉 En ARGOS lo más parecido es: **{sugerencias[0]}**" if sugerencias else "👉 (No se encontró nada parecido)"
                                alertas_nrc_rapido.append(f"❌ Buscó: **{llave_rota}** \n{sugerencia_txt}")
                            
                        df_c.insert(0, "NRC", nrc_asignados)
                        
                        df_c = df_c.rename(columns={
                            "TIPODEHORARIO": "TIPO DE HORARIO",
                            "METODO_EDUCATIVO": "METODO_ED"
                        })
                        
                        columnas_deseadas = ["NRC", "PERIODO", "SUBJ", "COURSE", "CAPACIDAD", "SECCION", "TIPO DE HORARIO", "METODO_ED", "datocomplementario"]
                        columnas_finales = [c for c in columnas_deseadas if c in df_c.columns]
                        
                        dfs_combinados.append(df_c[columnas_finales])
                        
                    if dfs_combinados:
                        df_resultado_rapido = pd.concat(dfs_combinados, ignore_index=True)
                        
                        # Guardamos el resultado completo
                        st.session_state.df_cruce_rapido = df_resultado_rapido
                        # Al generar un cruce nuevo, reseteamos la selección de columnas
                        st.session_state.pop("columnas_copia_rapida", None)
                        
                        st.success("✅ ¡Tabla cruzada generada exitosamente!")
                        
                        if alertas_nrc_rapido:
                            st.warning("⚠️ Ojo: Algunas filas no encontraron su NRC. Revisa las discrepancias abajo:")
                            with st.expander("🔍 Ver llaves que no cruzaron (Frente a Frente)"):
                                for a in alertas_nrc_rapido: st.write(a)
                    else:
                        st.error("No se pudo procesar la información de los CSV.")
                        
                except Exception as e:
                    st.error(f"❌ Ocurrió un error en el cruce rápido: {str(e)}")

        # ============================================================
        # RESULTADOS PARA COPIAR O DESCARGAR
        # ============================================================
        if st.session_state.df_cruce_rapido is not None:
            st.markdown("### 📋 Resultados del Cruce (NRC inyectados)")

            # Cambiamos solamente el nombre que VE el usuario.
            df_resultado_mostrar = st.session_state.df_cruce_rapido.rename(
                columns={"datocomplementario": "Cluster"}
            ).copy()

            # --------------------------------------------------------
            # SELECCIÓN DE COLUMNAS PARA COPIAR O DESCARGAR
            # --------------------------------------------------------
            columnas_seleccionadas = st.multiselect(
                "✅ Selecciona las columnas que quieres copiar o descargar:",
                options=list(df_resultado_mostrar.columns),
                default=list(df_resultado_mostrar.columns),
                key="columnas_copia_rapida",
                help="Puedes dejar todas o quitar las que no necesites."
            )

            if columnas_seleccionadas:
                df_para_copiar = df_resultado_mostrar[columnas_seleccionadas].copy()

                st.dataframe(
                    df_para_copiar,
                    use_container_width=True,
                    hide_index=True
                )

                col_b1, col_b2 = st.columns(2)

                # ----------------------------------------------------
                # COPIAR Y PEGAR DIRECTO EN EXCEL
                # ----------------------------------------------------
                with col_b1:
                    st.markdown("#### 📝 Copiar y pegar en Excel")
                    st.info(
                        "Haz clic en el botón de **Copiar** de la esquina "
                        "superior derecha del cuadro y después pégalo en Excel."
                    )
                    tsv_rapido = df_para_copiar.to_csv(index=False, sep="\t")
                    st.code(tsv_rapido, language="text")

                # ----------------------------------------------------
                # DESCARGAR EXCEL CON FORMATO Y COLORES
                # ----------------------------------------------------
                with col_b2:
                    st.markdown("#### 📥 Descargar en Excel")
                    st.info(
                        "El archivo descargado tendrá las columnas seleccionadas y formato visual listo."
                    )

                    # --- Aplicando openpyxl para darle formato a la descarga rápida ---
                    excel_rapido_buffer = io.BytesIO()
                    
                    with pd.ExcelWriter(excel_rapido_buffer, engine='openpyxl') as writer:
                        df_para_copiar.to_excel(writer, index=False, sheet_name="Cruce_NRC")
                        worksheet = writer.sheets["Cruce_NRC"]
                        
                        # Estilos de openpyxl
                        font_base = Font(name="Calibri", size=11)
                        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                        fill_header = PatternFill(start_color="1F4E78", fill_type="solid")
                        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        align_center = Alignment(horizontal="center", vertical="center")
                        
                        font_nrc = Font(name="Calibri", size=11, bold=True)
                        fill_nrc = PatternFill(start_color="DDEBF7", fill_type="solid")

                        # Colorear la Cabecera
                        for cell in worksheet[1]:
                            cell.font = font_header
                            cell.fill = fill_header
                            cell.alignment = align_header
                            
                        # Ajustar el ancho de las columnas
                        for col in worksheet.columns:
                            max_len = 0
                            for cell in col:
                                if cell.value:
                                    max_len = max(max_len, len(str(cell.value)))
                            worksheet.column_dimensions[col[0].column_letter].width = max(max_len + 3, 11)

                        # Detectar si se incluyó la columna NRC para resaltarla
                        nrc_col_idx = None
                        if "NRC" in df_para_copiar.columns:
                            nrc_col_idx = list(df_para_copiar.columns).index("NRC") + 1

                        # Aplicar fuente base, centrado y color al NRC
                        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                if nrc_col_idx and cell.column == nrc_col_idx:
                                    cell.font = font_nrc
                                    cell.fill = fill_nrc
                                else:
                                    cell.font = font_base
                                cell.alignment = align_center

                    excel_rapido_buffer.seek(0) # Rebobinamos el buffer para que Streamlit lo pueda leer

                    st.download_button(
                        label="📥 Descargar tabla formateada (.xlsx)",
                        data=excel_rapido_buffer.getvalue(),
                        file_name="Cruce_Rapido_NRC.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="dl_cruce_rapido_xlsx"
                    )

            else:
                st.warning(
                    "⚠️ Selecciona por lo menos una columna para poder copiar o descargar.")
